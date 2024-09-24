import logging
import io
import csv
from eventmanager import Evt
from data_export import DataExporter
from openpyxl import Workbook
from openpyxl.styles import Font

logger = logging.getLogger(__name__)


def assemble_last_race(rhapi):
    results = rhapi.eventresults.results

    if not results:
        logger.error("Unable to read results")
        return None

    payload = []

    event_name = rhapi.db.option("eventName")
    payload.append([event_name])

    last_class_id = max(results["classes"].keys())
    last_class_name = results["classes"][last_class_id]["name"]

    last_heat_id = max(results["heats"].keys())
    last_heat = results["heats"][last_heat_id]
    last_round = last_heat["rounds"][-1]

    payload.append(
        [
            f"{last_class_name} {last_heat['displayname']} {rhapi.__('Round')} {last_round['id']}"
        ]
    )

    laptimes = {}
    max_laps = 0

    for node in last_round["nodes"]:
        valid_laps = [lap for lap in node["laps"] if not lap["deleted"]]
        max_laps = max(max_laps, len(valid_laps))
        laptimes[node["callsign"]] = [lap["lap_time_formatted"] for lap in valid_laps]

    leaderboard_data = build_leaderboard(last_round["leaderboard"], rhapi)

    # Modify the header, add laps after the pilot column
    header = leaderboard_data[0]
    for i in range(0, max_laps):
        header.insert(
            2 + i, f"{rhapi.__('Lap')} {i}"
        )  # the index of Pilot is 2, so it is "2 + i"
    payload.append(header)

    # Add the lap time of each pilot
    for row in leaderboard_data[1:]:
        callsign = row[1]
        pilot_laptimes = laptimes.get(callsign, [])
        for i in range(max_laps):
            if i < len(pilot_laptimes):
                row.insert(2 + i, pilot_laptimes[i])
            else:
                row.insert(2 + i, "DNF")
        payload.append(row)

    return payload


def build_leaderboard(leaderboard, rhapi, **kwargs):
    if not leaderboard:
        return None

    meta = leaderboard["meta"]
    if "primary_leaderboard" in kwargs and kwargs["primary_leaderboard"] in leaderboard:
        primary_leaderboard = leaderboard[kwargs["primary_leaderboard"]]
    else:
        primary_leaderboard = leaderboard[meta["primary_leaderboard"]]

    if meta["start_behavior"] == 2:
        total_label = rhapi.__("Laps Total")
        total_source = "total_time_laps"
    else:
        total_label = rhapi.__("Total")
        total_source = "total_time"

    output = [
        [
            rhapi.__("Rank"),
            rhapi.__("Pilot"),
            rhapi.__("Fastest"),
            rhapi.__(total_label),
        ]
    ]

    for entry in primary_leaderboard:
        output.append(
            [
                entry["position"],
                entry["callsign"],
                entry["fastest_lap"],
                entry[total_source],
            ]
        )

    return output


def write_csv(data):
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_NONNUMERIC)
    writer.writerows(data)

    logger.info("Exporting to CSV")

    return {"data": output.getvalue(), "encoding": "text/csv", "ext": "csv"}


def write_excel(data):
    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)

    # Set font size for the first row
    for cell in ws[1]:
        cell.font = Font(size=16)

    # Set font size for the second row
    for cell in ws[2]:
        cell.font = Font(size=14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("Exporting to XLSX")

    return {
        "data": output.getvalue(),
        "encoding": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "ext": "xlsx",
    }


def register_handlers(args):
    for exporter in [
        DataExporter("Export Last Race(CSV)", write_csv, assemble_last_race),
        DataExporter("Export Last Race(XLSX)", write_excel, assemble_last_race),
    ]:
        args["register_fn"](exporter)


def initialize(rhapi):
    rhapi.events.on(Evt.DATA_EXPORT_INITIALIZE, register_handlers)
